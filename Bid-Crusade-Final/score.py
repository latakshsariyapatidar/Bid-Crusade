# ============================================
# BID A BIZ – TERMINAL SCORING ENGINE
# ============================================

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from collections import defaultdict
import os
import json

# -----------------------------
# CONFIGURATION
# -----------------------------

INITIAL_CAPITAL = 100.0
EXCEL_FILE = "Bid_Biz_Final_Score_Sheet.xlsx"
STATE_JSON_FILE = "bid_biz_state.json"

# -----------------------------
# ASSET DATABASE
# (ID → Name, Visible Category, Hidden Vertical, Base Price)
# Matching app.js database exactly (with dynamic JSON support)
# -----------------------------

default_assets = {
    1: ("Pan-India Distribution Rights", "A", "V3", 11),
    2: ("Multi-Specialty Hospital Unit", "A", "V1", 21),
    3: ("Multi-Sector ESG Certification Portfolio", "A", "V4", 6),
    4: ("Automated Robotics Assembly System", "A", "V2", 14),
    5: ("Regional Telecom Spectrum Allocation", "A", "V3", 18),
    6: ("Industrial Manufacturing Plant", "A", "V1", 21),
    7: ("Urban Smart Monitoring Network", "A", "V2", 9),
    8: ("Government Infrastructure Contract", "A", "V4", 16),
    9: ("Infrastructure Maintenance Rights", "A", "V4", 7),
    10: ("Mining & Raw Material Extraction Lease", "A", "V1", 22),
    11: ("National Logistics Fleet & Warehousing", "B", "V3", 15),
    12: ("Advanced Cybersecurity Suite", "B", "V2", 10),
    13: ("Tier-2 City Zoned Land Bank", "B", "V4", 15),
    14: ("Multi-Brand Automobile Assembly Unit", "B", "V1", 20),
    15: ("Export-Import Multi-Nation License", "B", "V3", 12),
    16: ("AI & Data Analytics Infrastructure", "B", "V2", 12),
    17: ("Private Power Backup Microgrid", "B", "V4", 9),
    18: ("Cross-Platform Mobile App Ecosystem", "B", "V2", 8),
    19: ("Cold Chain Pharma Distribution", "B", "V3", 13),
    20: ("Cloud Server Farm Facility", "B", "V1", 18),
    21: ("Regulatory Compliance Shield", "C", "V4", 8),
    22: ("Cross-Industry Patent Portfolio", "C", "V2", 9),
    23: ("Multi-State Retail Franchise Rights", "C", "V3", 13),
    24: ("Metropolitan Commercial Tower", "C", "V1", 19),
    25: ("Strategic Smart City Land Reserve", "C", "V4", 13),
    26: ("R&D Innovation Hub", "C", "V2", 11),
    27: ("Commercial Banking & NBFC License", "C", "V3", 14),
    28: ("Pharmaceutical Manufacturing Facility", "C", "V1", 19),
    29: ("Cold Storage Network", "C", "V3", 12),
    30: ("High-Speed Data Cable Network", "C", "V2", 11),
    31: ("Long-Term Raw Material Supply Contract", "D", "V4", 10),
    32: ("Agricultural Processing Facility", "D", "V1", 15),
    33: ("Scalable E-Commerce Platform", "D", "V3", 13),
    34: ("Data Center Cooling System", "D", "V2", 8),
    35: ("Commodity Trading Desk License", "D", "V4", 11),
    36: ("Green Hydrogen Pilot Plant", "D", "V1", 19),
    37: ("National Warehouse Network", "D", "V3", 15),
    38: ("Renewable Energy Solar Grid", "D", "V1", 20),
    39: ("Strategic Insurance & Risk Hedging", "D", "V4", 7),
    40: ("Enterprise ERP & Automation System", "D", "V2", 10)
}

assets = default_assets.copy()

# Dynamically load assets if state JSON exists
if os.path.exists(STATE_JSON_FILE):
    try:
        with open(STATE_JSON_FILE, "r") as f:
            data = json.load(f)
            if "assets" in data:
                dynamic_assets = {}
                for item in data["assets"]:
                    aid = int(item["id"])
                    name = item["name"]
                    category = item.get("category", "A")
                    vertical = item.get("vertical", "V1")
                    base_price = float(item.get("basePrice", 10))
                    dynamic_assets[aid] = (name, category, vertical, base_price)
                if dynamic_assets:
                    assets = dynamic_assets
                    print(f"Loaded {len(assets)} dynamic assets from {STATE_JSON_FILE}")
    except Exception as e:
        print(f"Using default asset database (JSON load error: {e})")

# -----------------------------
# TERMINAL INPUTS
# -----------------------------

team_name = input("Enter Team Name: ")

asset_input = input("Enter Asset IDs (comma separated): ")
asset_ids = [int(x.strip()) for x in asset_input.split(",") if x.strip().isdigit()]

remaining_budget = float(input("Enter Remaining Budget (₹ Cr): "))

disaster_code = input("Enter Disaster Code (W/X/Y/Z): ").upper()

# -----------------------------
# DISASTER LOGIC
# -----------------------------

def get_multiplier(category):
    if disaster_code == "W":
        if category == "A": return 1.20
        elif category == "D": return 0.80
        else: return 1.00

    if disaster_code == "X":
        if category == "B": return 1.20
        elif category == "A": return 0.80
        else: return 1.00

    if disaster_code == "Y":
        if category == "C": return 1.20
        elif category == "B": return 0.80
        else: return 1.00

    if disaster_code == "Z":
        if category == "D": return 1.20
        elif category == "C": return 0.80
        else: return 1.00

    return 1.00

# -----------------------------
# CALCULATIONS
# -----------------------------

total_investment = max(1.0, INITIAL_CAPITAL - remaining_budget)
post_value = 0.0

category_investment = defaultdict(float)
vertical_presence = {"V1":0, "V2":0, "V3":0, "V4":0}

for aid in asset_ids:
    if aid in assets:
        name, category, vertical, price = assets[aid]
        category_investment[category] += price
        vertical_presence[vertical] = 1
        post_value += price * get_multiplier(category)

roi = round((post_value - total_investment) / total_investment, 5)
max_exp_val = max(category_investment.values()) if category_investment else 0
max_exposure = round(max_exp_val / total_investment, 5)
vertical_count = sum(vertical_presence.values())

extra_asset = 1 if len(asset_ids) >= 5 else 0
roi_condition = 1 if roi >= 0.08 else 0
exposure_condition = 1 if max_exposure <= 0.45 else 0

vertical_score = (vertical_count / 4) * 15
risk_score = 15 if max_exposure <= 0.45 else round(15 * max(0, (1 - max_exposure) / 0.55) ** 2, 1)

roi_score = round(max(0, min(10, roi * 50)), 1)

business_strength = vertical_score + risk_score + roi_score

conditions = (1 if vertical_count == 4 and extra_asset else 0) + roi_condition + exposure_condition

if conditions == 3:
    bonus = 10
elif conditions == 2:
    bonus = 6
elif conditions == 1:
    bonus = 3
else:
    bonus = 0

organizer_total = business_strength + bonus

# -----------------------------
# PDF GENERATION
# -----------------------------

pdf_name = f"{team_name}_Summary.pdf"
doc = SimpleDocTemplate(pdf_name)
elements = []
styles = getSampleStyleSheet()

elements.append(Paragraph("<b>Bid A Biz Grand Finale – Organizer Summary</b>", styles["Title"]))
elements.append(Spacer(1, 0.3 * inch))

table_data = [
    ["Team Name", team_name],
    ["Initial Budget", INITIAL_CAPITAL],
    ["Total Investment", round(total_investment,2)],
    ["Post-Disaster Value", round(post_value,2)],
    ["ROI (%)", round(roi*100,2)],
    ["Exposure A (%)", round((category_investment["A"]/total_investment)*100 if "A" in category_investment else 0,2)],
    ["Exposure B (%)", round((category_investment["B"]/total_investment)*100 if "B" in category_investment else 0,2)],
    ["Exposure C (%)", round((category_investment["C"]/total_investment)*100 if "C" in category_investment else 0,2)],
    ["Exposure D (%)", round((category_investment["D"]/total_investment)*100 if "D" in category_investment else 0,2)],
    ["Max Exposure (%)", round(max_exposure*100,2)],
    ["V1 Present", vertical_presence["V1"]],
    ["V2 Present", vertical_presence["V2"]],
    ["V3 Present", vertical_presence["V3"]],
    ["V4 Present", vertical_presence["V4"]],
    ["Vertical Count", vertical_count],
    ["Extra Asset Condition", extra_asset],
    ["ROI ≥8%", roi_condition],
    ["Exposure ≤45%", exposure_condition],
    ["Business Strength (40)", round(business_strength,2)],
    ["Bonus (10)", bonus],
    ["Total Organizer Score (50)", round(organizer_total,2)],
]

table = Table(table_data, colWidths=[3*inch, 2.5*inch])
table.setStyle(TableStyle([("GRID", (0,0), (-1,-1), 1, colors.black)]))
elements.append(table)
doc.build(elements)

# -----------------------------
# EXCEL CREATE / APPEND
# -----------------------------

if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Score Sheet"
    headers = [
        "Team Name",
        "Business Strength (40)",
        "Bonus (10)",
        "Organizer Total (50)",
        "Strategic Depth (15)",
        "Market Logic (10)",
        "Clarity (10)",
        "Data & Disaster Explanation (10)",
        "Q&A Defense (5)",
        "Judge Total (50)",
        "Final Total (100)"
    ]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)

row = ws.max_row + 1

ws.cell(row=row, column=1, value=team_name)
ws.cell(row=row, column=2, value=round(business_strength,2))
ws.cell(row=row, column=3, value=bonus)
ws.cell(row=row, column=4, value=round(organizer_total,2))
ws.cell(row=row, column=10, value=f"=SUM(E{row}:I{row})")
ws.cell(row=row, column=11, value=f"=D{row}+J{row}")

wb.save(EXCEL_FILE)

print("\nCompleted.")
print("PDF Created:", pdf_name)
print("Excel Updated:", EXCEL_FILE)